import base64
import math
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from odoo import models, fields, api, _
import logging

_logger = logging.getLogger(__name__)


class StockSerialConversionWizard(models.TransientModel):
    _name = "stock.serial.conversion.wizard"
    _description = "Stock Serial Conversion Wizard"

    product_ids = fields.Many2many(
        "product.template", string="Products to Convert", readonly=True
    )
    state = fields.Selection(
        [("confirm", "Confirm"), ("done", "Done")], default="confirm"
    )
    processed_count = fields.Integer(readonly=True)
    ignored_count = fields.Integer(readonly=True)
    failed_count = fields.Integer(readonly=True)
    fractional_count = fields.Integer(readonly=True)
    failed_details = fields.Text(readonly=True)
    fractional_details = fields.Text(readonly=True)
    excel_file = fields.Binary(readonly=True, attachment=False)
    excel_filename = fields.Char(readonly=True)

    def _is_eligible(self, tmpl):
        if tmpl.type in ("service", "combo"):
            return False
        if hasattr(tmpl, "is_storable") and not tmpl.is_storable:
            return False
        if tmpl.tracking != "none":
            return False
        return True

    def _build_excel(self, excel_rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "Serial Number Adjustment"

        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="2E75B6")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, (header, width) in enumerate(
            zip(
                ["Location", "Serial Number", "Product", "Barcode", "Serial Number Database Id", "Counted"],
                [40, 18, 40, 25, 30, 12],
            ),
            1,
        ):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 20

        row_font = Font(name="Arial", size=10)
        alt_fill = PatternFill("solid", start_color="EBF3FB")

        for row_idx, row_data in enumerate(excel_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = row_font
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def action_convert(self):
        self.ensure_one()
        processed, ignored, failed, fractional_products, excel_rows = [], [], [], [], []

        Lot = self.env["stock.lot"]

        for tmpl in self.product_ids:
            if not self._is_eligible(tmpl):
                ignored.append(tmpl.name)
                continue
            try:
                with self.env.cr.savepoint():
                    serial_counter = 1

                    quants = self.env["stock.quant"].search(
                        [
                            ("product_id.product_tmpl_id", "=", tmpl.id),
                            ("location_id.usage", "=", "internal"),
                            ("quantity", ">", 0),
                        ]
                    )

                    has_fractional = False
                    quant_data = []

                    for quant in quants:
                        floored = math.floor(quant.quantity)
                        if abs(quant.quantity - floored) > 1e-9:
                            has_fractional = True
                        if floored > 0:
                            quant_data.append((quant, floored))

                    if quant_data:
                        for quant, _qty in quant_data:
                            quant.with_context(inventory_mode=True).sudo().write(
                                {
                                    "inventory_quantity": 0,
                                    "inventory_quantity_set": True,
                                    "inventory_date": datetime.now(),
                                    "user_id": self.env.uid,
                                }
                            )

                    for quant, floored_qty in quant_data:
                        product = quant.product_id

                        for idx in range(floored_qty):
                            serial_name = str(serial_counter)

                            existing_lot = Lot.search(
                                [
                                    ("name", "=", serial_name),
                                    ("product_id", "=", product.id),
                                ],
                                limit=1,
                            )

                            lot = existing_lot
                            if not existing_lot:
                                lot = Lot.create(
                                    {
                                        "name": serial_name,
                                        "product_id": product.id,
                                        "company_id": self.env.company.id,
                                    }
                                )

                            excel_rows.append(
                                [
                                    quant.location_id.complete_name,
                                    serial_name,
                                    tmpl.name,
                                    product.barcode or "",
                                    lot.id,
                                    1,
                                ]
                            )

                            serial_counter += 1

                    tmpl.tracking = "serial"
                    tmpl.message_post(
                        body=(
                            "Tracking converted to Serial Number by %s on %s "
                            "via the Serial Number Conversion wizard."
                        )
                        % (
                            self.env.user.name,
                            fields.Datetime.now(),
                        )
                    )

                    if has_fractional:
                        fractional_products.append(tmpl.name)

                    processed.append(tmpl.name)

            except Exception as e:
                failed.append("%s: %s" % (tmpl.name, str(e)))

        excel_data, filename = False, False
        if excel_rows:
            excel_data = base64.b64encode(self._build_excel(excel_rows))
            filename = "serial_number_adjustment.xlsx"

        self.write(
            {
                "state": "done",
                "processed_count": len(processed),
                "ignored_count": len(ignored),
                "failed_count": len(failed),
                "fractional_count": len(fractional_products),
                "failed_details": "\n".join(failed) or False,
                "fractional_details": "\n".join(fractional_products) or False,
                "excel_file": excel_data,
                "excel_filename": filename,
            }
        )
        
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': excel_data,
            'res_model': 'res.company',        
            'res_id': self.env.company.id,     
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }




